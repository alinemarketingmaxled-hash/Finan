#!/usr/bin/env python3
"""
ETL: DRE_MAX_LED_2026.xlsx -> JSON data files for the web dashboard.

Reads the raw transaction sheets (source of truth) from SRC and a handful of
pre-aggregated helper cells from the "Home" sheet (2025 history, empréstimos,
a receber/pagar) which don't exist as raw transactions and produces
normalized JSON under data/generated/.

SRC vs HOME_SRC: a workbook update replaced the transaction sheets (added
junho/2026 and beyond) but dropped the "Home" sheet entirely. Home's content
didn't change between the previous and current transactional files, so we
keep reading it from the last file that still has it (HOME_SRC) instead of
losing 2025 history/empréstimos/a receber-pagar. If a future update brings
Home back into the main file, point HOME_SRC back at SRC.
"""
import json
import re
import unicodedata
from collections import defaultdict
from datetime import datetime
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent.parent
SRC = ROOT / "data" / "source" / "DRE_MAX_LED_2026.xlsx"
HOME_SRC = ROOT / "data" / "source" / "DRE_MAX_LED_2026_home_ref.xlsx"
OUT = ROOT / "data" / "generated"
OUT.mkdir(parents=True, exist_ok=True)
SITE_JS = ROOT / "site" / "js"
SITE_JS.mkdir(parents=True, exist_ok=True)

MONTH_PT = ["Jan", "Fev", "Mar", "Abr", "Mai", "Jun", "Jul", "Ago", "Set", "Out", "Nov", "Dez"]

wb = openpyxl.load_workbook(SRC, data_only=True)


def mkey(d: datetime) -> str:
    return f"{d.year:04d}-{d.month:02d}"


PT_MONTH_ABBR = {"jan": 1, "fev": 2, "mar": 3, "abr": 4, "mai": 5, "jun": 6,
                 "jul": 7, "ago": 8, "set": 9, "out": 10, "nov": 11, "dez": 12}


def parse_month_cell(v):
    """Normalizes a month cell that may be a datetime OR a 'mmm/yy' string
    (the source sheet mixes both in the same column) into a 'YYYY-MM' key."""
    if isinstance(v, datetime):
        return mkey(v)
    if isinstance(v, str):
        m = re.match(r"([a-zçõ]+)/(\d{2,4})", v.strip().lower())
        if m:
            abbr, yy = m.groups()
            mm = PT_MONTH_ABBR.get(abbr[:3])
            if mm:
                year = int(yy)
                if year < 100:
                    year += 2000
                return f"{year:04d}-{mm:02d}"
    return None


def norm_cat(raw):
    if raw is None:
        return None
    s = str(raw).strip().upper()
    s = re.sub(r"\s+", " ", s)
    s = unicodedata.normalize("NFC", s)
    aliases = {
        "EMPRÉSTIMOS FGI": "EMPRÉSTIMO FGI",
        "IMPORTAÇÕES": "IMPORTAÇÃO",
        "SEGURO SÓCIOS CARRO": "SEGURO SÓCIO CARRO",
        "OUTROS": "OUTRAS DESPESAS",
        # variações/erros de digitação que apareceram na atualização com junho/2026
        "EMPRESTIMO": "EMPRÉSTIMO",
        "EMPRESTIMOS": "EMPRÉSTIMO",
        "EMPRÉSTIMOS": "EMPRÉSTIMO",
        "IMPOSTO": "IMPOSTOS",
        "GASTOS FICOS": "GASTOS FIXOS",
        "INSUMO": "INSUMOS",
        "SÓCIO": "SÓCIOS",
        "SEGURO": "SEGURO SAÚDE",
        "ANUAL": "OUTRAS DESPESAS",
    }
    return aliases.get(s, s)


GROUPS = {
    "FORNECEDORES": "Custo de Mercadorias",
    "IMPOSTOS": "Impostos",
    "BANCO": "Financeiro/Bancário",
    "CARTÃO DE CRÉDITO": "Financeiro/Bancário",
    "CONTABILIDADE": "Financeiro/Bancário",
    "MOVIMENTO": "Financeiro/Bancário",
    "MATERIAL DE ESCRITÓRIO": "Administrativo",
    "GASTOS FIXOS": "Administrativo",
    "MANUTENÇÃO": "Administrativo",
    "CONSUMO": "Administrativo",
    "INVESTIMENTO": "Administrativo",
    "FUNCIONÁRIOS": "Pessoal",
    "FUNCIONÁRIOS (VR)": "Pessoal",
    "SÓCIOS": "Pessoal",
    "PLANO DE SAÚDE": "Pessoal",
    "SEGURO DE VIDA SÓCIO": "Pessoal",
    "SEGURO SÓCIO CARRO": "Pessoal",
    "EMPRÉSTIMO CAPITAL DE GIRO BRADESCO": "Dívida/Empréstimos",
    "EMPRÉSTIMO CAPITAL DE GIRO ITAÚ": "Dívida/Empréstimos",
    "EMPRÉSTIMO PRONAMP": "Dívida/Empréstimos",
    "EMPRÉSTIMO FGI": "Dívida/Empréstimos",
    "SEGURO DE EMPRÉSTIMO": "Dívida/Empréstimos",
    "MARKETING": "Marketing & Vendas",
    "PUBLICIDADE": "Marketing & Vendas",
    "BRINDES": "Marketing & Vendas",
    "IMPORTAÇÃO": "Logística/Importação",
    "TRANSPORTES": "Logística/Importação",
    "TRANSPORTES IMPORTAÇÃO": "Logística/Importação",
    "DEVOLUÇÃO": "Outras",
    "INSUMOS": "Outras",
    "SERVIÇOS": "Outras",
    "OUTRAS DESPESAS": "Outras",
    "EMPRÉSTIMO": "Dívida/Empréstimos",
    "SEGURO SAÚDE": "Pessoal",
}

DIVISIONS = ["iluminacao", "importacao"]
DIVISION_LABELS = {
    "iluminacao": "Max Led Iluminação",
    "importacao": "Max Led Importação",
    "consolidado": "Max Led Consolidado",
}

# ---------------------------------------------------------------------------
# 1. Raw transactions
# ---------------------------------------------------------------------------
tx = []  # list of dicts: date, division, basis, flow, category, counterparty, value


def fmt_nota(v):
    """Normaliza o número/referência da nota fiscal (célula pode ser int, float ou string)."""
    if v is None or v == "":
        return None
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v).strip() or None


def read_rows(sheet, division, basis, flow, date_col, cat_col, cp_col, val_col, nota_col=None):
    ws = wb[sheet]
    for row in ws.iter_rows(min_row=2):
        d = row[date_col].value
        v = row[val_col].value
        if not isinstance(d, datetime) or not isinstance(v, (int, float)):
            continue
        cat = norm_cat(row[cat_col].value) if cat_col is not None else None
        cp = row[cp_col].value if cp_col is not None else None
        nota = fmt_nota(row[nota_col].value) if nota_col is not None else None
        tx.append(
            dict(
                date=d, division=division, basis=basis, flow=flow,
                category=cat, counterparty=(str(cp).strip() if cp else None),
                value=float(v), nota_fiscal=nota,
            )
        )


# Iluminação - financeiro (cash). "Nota" é a referência da nota fiscal (col. 2).
read_rows("ENTRADAS - FIN - ilumi", "iluminacao", "financeiro", "entrada", 0, None, 1, 3, nota_col=2)
read_rows("SAIDAS-FIN-ilumi", "iluminacao", "financeiro", "saida", 0, 2, 1, 3)
# Iluminação - NFe (invoice). Nessa aba os cabeçalhos "Cliente"/"Nota" estão trocados
# em relação ao conteúdo real: col. 1 é o número da nota, col. 2 é o nome do cliente.
read_rows("ENTRADA NFE ILUMI", "iluminacao", "nfe", "compra", 0, None, 2, 3, nota_col=1)
read_rows("SAÍDA NFE ILUMI", "iluminacao", "nfe", "venda", 0, None, 2, 3, nota_col=1)
# Importação - financeiro (cash). "Justificativa" nas entradas é na prática a
# referência da nota/pedido (numérica na maioria das linhas).
read_rows("Entradas- fin- import", "importacao", "financeiro", "entrada", 0, None, 1, 3, nota_col=2)
read_rows("Saidas-fin-import", "importacao", "financeiro", "saida", 0, 2, 1, 3)
# Importação - NFe (invoice)
read_rows("Entradas-NFE-Import", "importacao", "nfe", "compra", 0, None, 2, 3, nota_col=1)
read_rows("Saídas-NFE-import", "importacao", "nfe", "venda", 0, None, 2, 3, nota_col=1)

print(f"Total transactions parsed: {len(tx)}")
detailed_months = sorted({mkey(t["date"]) for t in tx})
print(f"Detailed months present: {detailed_months}")

# Guarda-corpo: categoria que não bate com nenhum grupo conhecido (nem direto,
# nem via ALIASES) cai silenciosamente em "Outras" no DRE -- imprime pra nunca
# passar batido num pagamento de imposto/fornecedor mal categorizado.
unknown_cats = defaultdict(lambda: [0, 0.0])
for t in tx:
    if t["category"] and t["category"] not in GROUPS:
        unknown_cats[t["category"]][0] += 1
        unknown_cats[t["category"]][1] += t["value"]
if unknown_cats:
    print(f"\n*** ATENÇÃO: {len(unknown_cats)} categoria(s) não reconhecida(s), caindo em 'Outras' -- considere adicionar em ALIASES/GROUPS: ***")
    for cat, (n, total) in sorted(unknown_cats.items(), key=lambda kv: -kv[1][1]):
        print(f"    {cat!r}: {n} lançamento(s), R$ {total:,.2f}")
    print()

# Full transaction export (powers the Lançamentos ledger page + lets every
# other view be recomputed client-side once manual entries are merged in).
tx_sorted = sorted(tx, key=lambda t: t["date"])
tx_export = [
    dict(
        id=f"tx{i:05d}", date=t["date"].date().isoformat(), division=t["division"],
        basis=t["basis"], flow=t["flow"], category=t["category"],
        counterparty=t["counterparty"], value=round(t["value"], 2),
        nota_fiscal=t.get("nota_fiscal"),
    )
    for i, t in enumerate(tx_sorted)
]
(OUT / "transactions.json").write_text(json.dumps(tx_export, ensure_ascii=False, indent=2))
print(f"transactions.json: {len(tx_export)} rows")

# ---------------------------------------------------------------------------
# 2. Monthly cashflow (financeiro basis): 2025 history (pre-aggregated,
#    hardcoded in the workbook) + 2026 detailed (from tx) + Jun-Sep/2026
#    forecast (workbook's own FORECAST formulas, already evaluated).
# ---------------------------------------------------------------------------
home = openpyxl.load_workbook(HOME_SRC, data_only=True)["Home"]

cashflow = []

# 2025 actuals per division, hardcoded monthly in Home!B56:D67 (ilu) and H56:J67 (imp)
for r in range(56, 68):
    d = home.cell(row=r, column=2).value  # B
    ent_i = home.cell(row=r, column=3).value  # C
    sai_i = home.cell(row=r, column=4).value  # D
    ent_m = home.cell(row=r, column=9).value  # I
    sai_m = home.cell(row=r, column=10).value  # J
    if not isinstance(d, datetime):
        continue
    cashflow.append(dict(month=mkey(d), division="iluminacao", entradas=ent_i, saidas=sai_i,
                          resultado=ent_i - sai_i, tipo="realizado"))
    cashflow.append(dict(month=mkey(d), division="importacao", entradas=ent_m, saidas=sai_m,
                          resultado=ent_m - sai_m, tipo="realizado"))

# 2026-01 .. last detailed month: derive from transactions (ground truth)
agg = defaultdict(lambda: defaultdict(float))
for t in tx:
    if t["basis"] != "financeiro":
        continue
    k = (mkey(t["date"]), t["division"])
    agg[k]["entradas" if t["flow"] == "entrada" else "saidas"] += t["value"]

for (month, division), vals in agg.items():
    ent = vals.get("entradas", 0.0)
    sai = vals.get("saidas", 0.0)
    cashflow.append(dict(month=month, division=division, entradas=round(ent, 2),
                          saidas=round(sai, 2), resultado=round(ent - sai, 2), tipo="realizado"))

# Forecast Jun-Sep/2026: Home!C72:E76 (ilu) / I72:K76 (imp), skip stale row 72 (May,
# already realized above) and only take rows for months AFTER the last detailed month.
last_detailed = max(detailed_months)
for r in range(72, 77):
    d = home.cell(row=r, column=2).value  # B (date anchor for ilu block)
    ent_i = home.cell(row=r, column=3).value
    sai_i = home.cell(row=r, column=4).value
    ent_m = home.cell(row=r, column=9).value
    sai_m = home.cell(row=r, column=10).value
    if not isinstance(d, datetime):
        continue
    mk = mkey(d)
    if mk <= last_detailed:
        continue
    cashflow.append(dict(month=mk, division="iluminacao", entradas=round(ent_i, 2), saidas=round(sai_i, 2),
                          resultado=round(ent_i - sai_i, 2), tipo="previsao"))
    cashflow.append(dict(month=mk, division="importacao", entradas=round(ent_m, 2), saidas=round(sai_m, 2),
                          resultado=round(ent_m - sai_m, 2), tipo="previsao"))

# Consolidado = sum of both divisions per month/tipo
by_month = defaultdict(lambda: defaultdict(float))
tipo_by_month = {}
for row in cashflow:
    by_month[row["month"]]["entradas"] += row["entradas"]
    by_month[row["month"]]["saidas"] += row["saidas"]
    tipo_by_month[row["month"]] = row["tipo"]
for month, vals in by_month.items():
    cashflow.append(dict(month=month, division="consolidado", entradas=round(vals["entradas"], 2),
                          saidas=round(vals["saidas"], 2),
                          resultado=round(vals["entradas"] - vals["saidas"], 2),
                          tipo=tipo_by_month[month]))

cashflow.sort(key=lambda r: (r["month"], r["division"]))
(OUT / "monthly_cashflow.json").write_text(json.dumps(cashflow, ensure_ascii=False, indent=2))
print(f"monthly_cashflow.json: {len(cashflow)} rows")

# ---------------------------------------------------------------------------
# 3. Monthly NFe (invoice) basis - detailed months only
# ---------------------------------------------------------------------------
agg_nfe = defaultdict(lambda: defaultdict(float))
for t in tx:
    if t["basis"] != "nfe":
        continue
    k = (mkey(t["date"]), t["division"])
    agg_nfe[k]["vendas" if t["flow"] == "venda" else "compras"] += t["value"]

nfe_rows = []
for (month, division), vals in agg_nfe.items():
    vendas = round(vals.get("vendas", 0.0), 2)
    compras = round(vals.get("compras", 0.0), 2)
    nfe_rows.append(dict(month=month, division=division, vendas=vendas, compras=compras,
                          resultado=round(vendas - compras, 2)))

by_month_nfe = defaultdict(lambda: defaultdict(float))
for row in nfe_rows:
    by_month_nfe[row["month"]]["vendas"] += row["vendas"]
    by_month_nfe[row["month"]]["compras"] += row["compras"]
for month, vals in by_month_nfe.items():
    nfe_rows.append(dict(month=month, division="consolidado", vendas=round(vals["vendas"], 2),
                          compras=round(vals["compras"], 2),
                          resultado=round(vals["vendas"] - vals["compras"], 2)))

nfe_rows.sort(key=lambda r: (r["month"], r["division"]))
(OUT / "monthly_nfe.json").write_text(json.dumps(nfe_rows, ensure_ascii=False, indent=2))
print(f"monthly_nfe.json: {len(nfe_rows)} rows")

# ---------------------------------------------------------------------------
# 4. DRE mensal (financeiro) por divisao, com despesas por categoria
# ---------------------------------------------------------------------------
def build_dre(month, division):
    def flt(basis, flow, cat=None):
        s = 0.0
        for t in tx:
            if t["basis"] != basis or t["flow"] != flow:
                continue
            if division != "consolidado" and t["division"] != division:
                continue
            if mkey(t["date"]) != month:
                continue
            if cat is not None and t["category"] != cat:
                continue
            s += t["value"]
        return s

    receita_bruta = flt("financeiro", "entrada")
    impostos = flt("financeiro", "saida", "IMPOSTOS")
    receita_liquida = receita_bruta - impostos
    custo_mercadorias = flt("financeiro", "saida", "FORNECEDORES")
    lucro_bruto = receita_liquida - custo_mercadorias

    # all expense categories except FORNECEDORES/IMPOSTOS (already used above)
    cats = defaultdict(float)
    for t in tx:
        if t["basis"] != "financeiro" or t["flow"] != "saida":
            continue
        if division != "consolidado" and t["division"] != division:
            continue
        if mkey(t["date"]) != month:
            continue
        if t["category"] in ("IMPOSTOS", "FORNECEDORES"):
            continue
        cats[t["category"]] += t["value"]

    despesas = [
        dict(categoria=c, grupo=GROUPS.get(c, "Outras"), valor=round(v, 2))
        for c, v in sorted(cats.items(), key=lambda kv: -kv[1])
    ]
    despesas_total = sum(d["valor"] for d in despesas)
    resultado_operacional = lucro_bruto - despesas_total

    return dict(
        month=month, division=division, basis="financeiro",
        receita_bruta=round(receita_bruta, 2), impostos=round(impostos, 2),
        receita_liquida=round(receita_liquida, 2), custo_mercadorias=round(custo_mercadorias, 2),
        lucro_bruto=round(lucro_bruto, 2), despesas=despesas, despesas_total=round(despesas_total, 2),
        resultado_operacional=round(resultado_operacional, 2),
        margem_bruta=round(lucro_bruto / receita_liquida, 4) if receita_liquida else 0,
        margem_liquida=round(resultado_operacional / receita_bruta, 4) if receita_bruta else 0,
    )


dre_rows = []
for month in detailed_months:
    for division in DIVISIONS + ["consolidado"]:
        dre_rows.append(build_dre(month, division))

(OUT / "dre_monthly.json").write_text(json.dumps(dre_rows, ensure_ascii=False, indent=2))
print(f"dre_monthly.json: {len(dre_rows)} rows")

# ---------------------------------------------------------------------------
# 5. Expense categories aggregated over the whole detailed window
# ---------------------------------------------------------------------------
def expense_categories(division):
    cats = defaultdict(lambda: defaultdict(float))
    receita_total = 0.0
    for t in tx:
        if division != "consolidado" and t["division"] != division:
            continue
        if t["basis"] != "financeiro":
            continue
        if t["flow"] == "entrada":
            receita_total += t["value"]
            continue
        if t["category"] in ("IMPOSTOS", "FORNECEDORES"):
            continue
        cats[t["category"]]["total"] += t["value"]
        cats[t["category"]][mkey(t["date"])] += t["value"]

    out = []
    for c, vals in cats.items():
        total = vals.pop("total")
        out.append(dict(
            division=division, categoria=c, grupo=GROUPS.get(c, "Outras"),
            valor=round(total, 2),
            pct_receita=round(total / receita_total, 4) if receita_total else 0,
            meses={k: round(v, 2) for k, v in vals.items()},
        ))
    out.sort(key=lambda d: -d["valor"])
    return out


expense_cats = []
for division in DIVISIONS + ["consolidado"]:
    expense_cats.extend(expense_categories(division))
(OUT / "expense_categories.json").write_text(json.dumps(expense_cats, ensure_ascii=False, indent=2))
print(f"expense_categories.json: {len(expense_cats)} rows")

# ---------------------------------------------------------------------------
# 6. Counterparties (top suppliers / clients)
# ---------------------------------------------------------------------------
def top_counterparties(division, flow, basis, tipo_label, n=12):
    agg = defaultdict(lambda: [0.0, 0])
    for t in tx:
        if division != "consolidado" and t["division"] != division:
            continue
        if t["flow"] != flow or t["basis"] != basis or not t["counterparty"]:
            continue
        agg[t["counterparty"]][0] += t["value"]
        agg[t["counterparty"]][1] += 1
    rows = [dict(division=division, tipo=tipo_label, basis=basis, nome=k, valor=round(v[0], 2), n_transacoes=v[1])
            for k, v in agg.items()]
    rows.sort(key=lambda r: -r["valor"])
    return rows[:n]


counterparties = []
for division in DIVISIONS + ["consolidado"]:
    counterparties.extend(top_counterparties(division, "entrada", "financeiro", "cliente"))
    counterparties.extend(top_counterparties(division, "saida", "financeiro", "fornecedor"))
(OUT / "counterparties.json").write_text(json.dumps(counterparties, ensure_ascii=False, indent=2))
print(f"counterparties.json: {len(counterparties)} rows")

# ---------------------------------------------------------------------------
# 7. Loans
# ---------------------------------------------------------------------------
def d2s(v):
    return v.date().isoformat() if isinstance(v, datetime) else v


loans = [
    dict(id="bradesco_giro", nome="Capital de Giro Bradesco", divisao="iluminacao", banco="Bradesco",
         valor_total=home["D110"].value, parcelas_total=home["D111"].value, parcelas_pagas=home["D112"].value,
         parcela_com_juros=round(home["D113"].value, 2), parcelas_restantes=home["D114"].value,
         valor_restante=round(home["D115"].value, 2), valor_pago=round(home["D116"].value, 2),
         acrescimo=round(home["D117"].value, 2), valor_final_com_acrescimo=round(home["D118"].value, 2),
         data_inicial=d2s(home["D119"].value), data_final_prevista=d2s(home["D120"].value)),
    dict(id="itau_giro", nome="Capital de Giro Itaú", divisao="iluminacao", banco="Itaú",
         valor_total=home["F110"].value, parcelas_total=home["F111"].value, parcelas_pagas=home["F112"].value,
         parcela_com_juros=round(home["F113"].value, 2), parcelas_restantes=home["F114"].value,
         valor_restante=round(home["F115"].value, 2), valor_pago=round(home["F116"].value, 2),
         acrescimo=round(home["F117"].value, 2), valor_final_com_acrescimo=round(home["F118"].value, 2),
         data_inicial=d2s(home["F119"].value), data_final_prevista=d2s(home["F120"].value)),
    dict(id="pronamp", nome="Empréstimo Pronamp", divisao="importacao", banco="Pronamp",
         valor_total=home["J110"].value, parcelas_total=home["J111"].value, parcelas_pagas=home["J112"].value,
         parcela_com_juros=round(home["J113"].value, 2), parcelas_restantes=home["J114"].value,
         valor_restante=round(home["J116"].value, 2), valor_pago=round(home["J115"].value, 2),
         acrescimo=round(home["J117"].value, 2), valor_final_com_acrescimo=round(home["J118"].value, 2),
         data_inicial=d2s(home["J119"].value), data_final_prevista=d2s(home["J120"].value)),
    dict(id="fgi", nome="Capital de Giro FGI", divisao="importacao", banco="FGI",
         valor_total=home["L110"].value, parcelas_total=home["L111"].value, parcelas_pagas=home["L112"].value,
         parcela_com_juros=round(home["L113"].value, 2), parcelas_restantes=home["L114"].value,
         valor_restante=round(home["L116"].value, 2), valor_pago=round(home["L115"].value, 2),
         acrescimo=round(home["L117"].value, 2), valor_final_com_acrescimo=round(home["L118"].value, 2),
         data_inicial=d2s(home["L119"].value), data_final_prevista=d2s(home["L120"].value)),
]
(OUT / "loans.json").write_text(json.dumps(loans, ensure_ascii=False, indent=2))
print(f"loans.json: {len(loans)} rows")

# ---------------------------------------------------------------------------
# 8. Receivables / payables (A Receber / A Pagar) monthly
# ---------------------------------------------------------------------------
recv = []
for r in range(90, 95):
    m = parse_month_cell(home.cell(row=r, column=2).value)  # B
    a_receber_i = home.cell(row=r, column=3).value
    a_pagar_i = home.cell(row=r, column=4).value
    m2 = parse_month_cell(home.cell(row=r, column=8).value)  # H
    a_receber_m = home.cell(row=r, column=9).value
    a_pagar_m = home.cell(row=r, column=10).value
    if m:
        recv.append(dict(month=m, division="iluminacao",
                          a_receber=a_receber_i, a_pagar=a_pagar_i, saldo=round(a_receber_i - a_pagar_i, 2)))
    if m2:
        recv.append(dict(month=m2, division="importacao",
                          a_receber=a_receber_m, a_pagar=a_pagar_m, saldo=round(a_receber_m - a_pagar_m, 2)))
recv.sort(key=lambda r: (r["month"], r["division"]))
(OUT / "receivables_payables.json").write_text(json.dumps(recv, ensure_ascii=False, indent=2))
print(f"receivables_payables.json: {len(recv)} rows")

# ---------------------------------------------------------------------------
# 9. Meta
# ---------------------------------------------------------------------------
meta = dict(
    generated_at=datetime.utcnow().isoformat() + "Z",
    company="Max Led",
    divisions=DIVISIONS,
    division_labels=DIVISION_LABELS,
    detailed_range=dict(start=min(detailed_months), end=max(detailed_months)),
    history_range=dict(start="2025-01", end=max(detailed_months)),
    forecast_range=dict(start="2026-06", end="2026-09"),
    currency="BRL",
)
(OUT / "meta.json").write_text(json.dumps(meta, ensure_ascii=False, indent=2))
print("meta.json written")

# ---------------------------------------------------------------------------
# 10. Bundle everything into a single classic-script data file for the site
#     (no bundler / no ES modules, so it works straight off the filesystem).
# ---------------------------------------------------------------------------
bundle = dict(
    meta=meta,
    transactions=tx_export,
    cashflow=cashflow,
    nfe=nfe_rows,
    dre=dre_rows,
    expenseCategories=expense_cats,
    counterparties=counterparties,
    loans=loans,
    receivablesPayables=recv,
)
js_out = "// AUTO-GERADO por scripts/extract_xlsx.py -- não editar à mão.\n"
js_out += f"const MAXLED_DATA = {json.dumps(bundle, ensure_ascii=False, indent=2)};\n"
(SITE_JS / "data.js").write_text(js_out, encoding="utf-8")
print(f"site/js/data.js written ({len(js_out):,} bytes)")

print("DONE")
